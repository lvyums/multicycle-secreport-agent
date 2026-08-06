"""Excel 适配器 — 解析威胁情报台账 xlsx（openpyxl，V1.2）"""

import os
from typing import Optional

from capability.adapter.adapter_base import DataSourceAdapter
from common.logger.logger import LogManager

logger = LogManager.get_logger()

# 置信度 → 风险初判
CONFIDENCE_RISK = {"高": "MEDIUM", "中": "LOW", "低": "INFO", "HIGH": "MEDIUM",
                   "MEDIUM": "LOW", "LOW": "INFO"}


class ExcelAdapter(DataSourceAdapter):
    """Excel 导入适配器（威胁情报台账/资产清单场景）"""

    type = "EXCEL"

    def validate_config(self) -> list[str]:
        errors = []
        file_path = (self.config.config_json or {}).get("file_path", "")
        if not file_path:
            errors.append("缺少 file_path 配置")
        elif not os.path.exists(file_path):
            errors.append(f"Excel 文件不存在: {file_path}")
        return errors

    def fetch(self, window_start: str, window_end: str, task_id: int = 0) -> list[dict]:
        file_path = (self.config.config_json or {}).get("file_path", "")
        if not os.path.exists(file_path):
            logger.error(f"[EXCEL] 文件不存在: {file_path}")
            return []
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("[EXCEL] openpyxl 未安装，无法解析 xlsx")
            return []

        events: list[dict] = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            logger.error(f"[EXCEL] 解析失败: {e}")
            return []
        if not rows:
            return []

        header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            item = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
            parsed = self.parse_row(item, window_start, window_end)
            if parsed:
                events.append(parsed)
        logger.info(f"[EXCEL] {self.name} 拉取 {len(events)} 条（窗口 {window_start}~{window_end}）")
        return events

    def parse_row(self, item: dict, window_start: str, window_end: str) -> Optional[dict]:
        """解析一行情报 → raw dict（发布时间窗口过滤）"""
        ts = str(item.get("发布时间") or "").strip()
        if not ts or ts < window_start or ts > window_end:
            return None
        confidence = str(item.get("置信度") or "中").strip()
        risk = CONFIDENCE_RISK.get(confidence, "LOW")
        return {
            "source_type": "EXCEL",
            "source_name": self.name,
            "receive_time": ts,
            "raw_content": str(item),
            "status": "OK",
            "extra": {
                "event_type": "threat_intel",
                "risk_hint": risk,
                "asset_ip": str(item.get("影响资产") or "").strip(),
                "device": "threat-intel",
                "intel_name": str(item.get("情报名称") or "").strip(),
                "intel_type": str(item.get("情报类型") or "").strip(),
                "confidence": confidence,
                "source": str(item.get("来源") or "").strip(),
                "suggestion": str(item.get("处置建议") or "").strip(),
            },
        }
